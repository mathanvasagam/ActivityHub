import csv
from datetime import timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook

from .forms import BlogPostForm, LinkedInPostForm, PostFilterForm, ProjectForm, ResearchPaperForm, UserProfileForm
from .models import BlogPost, LinkedInPost, Project, ResearchPaper, UserProfile


def _current_mode(request) -> str:
    if not request.user.is_authenticated:
        return "guest"
    return "admin" if request.user.is_staff else "user"


def _apply_post_filters(queryset, form: PostFilterForm):
    if not form.is_valid():
        return queryset

    period = form.cleaned_data.get("period") or "all"
    q = form.cleaned_data.get("q")
    company_name = form.cleaned_data.get("company_name")
    user_name = form.cleaned_data.get("user_name")
    start_date = form.cleaned_data.get("start_date")
    end_date = form.cleaned_data.get("end_date")

    if period == "last_7_days":
        queryset = queryset.filter(posted_at__date__gte=timezone.localdate() - timedelta(days=7))
    elif period == "last_30_days":
        queryset = queryset.filter(posted_at__date__gte=timezone.localdate() - timedelta(days=30))
    elif period == "custom":
        if start_date:
            queryset = queryset.filter(posted_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(posted_at__date__lte=end_date)

    if company_name:
        queryset = queryset.filter(company_name__icontains=company_name)
    if user_name:
        queryset = queryset.filter(user__username__icontains=user_name)
    if q:
        queryset = queryset.filter(
            Q(post_title__icontains=q)
            | Q(company_name__icontains=q)
            | Q(notes__icontains=q)
        )

    return queryset


def _require_user_mode(request):
    if request.user.is_staff:
        messages.error(request, "Admin accounts can only access the organization dashboard.")
        return redirect("searcher:organization")
    return None


def landing_view(request):
    if request.user.is_authenticated:
        return redirect("searcher:posts")
    return render(request, "searcher/landing.html")


@login_required
def home_view(request):
    if request.user.is_staff and _current_mode(request) == "admin":
        return redirect("searcher:organization")

    filter_form = PostFilterForm(request.GET or None)
    user_posts = LinkedInPost.objects.filter(user=request.user)
    user_posts = _apply_post_filters(user_posts, filter_form)

    return render(
        request,
        "searcher/home.html",
        {
            "total_posts": user_posts.count(),
            "last_7_days": user_posts.filter(posted_at__date__gte=timezone.localdate() - timedelta(days=7)).count(),
            "total_projects": Project.objects.filter(user=request.user).count(),
            "total_blogs": BlogPost.objects.filter(user=request.user).count(),
            "total_research_papers": ResearchPaper.objects.filter(user=request.user).count(),
            "posts": user_posts,
            "filter_form": filter_form,
        },
    )


@login_required
def create_options_view(request):
    if request.user.is_staff:
        return JsonResponse({"ok": False, "error": "Admin accounts cannot create user content."}, status=403)

    return JsonResponse(
        {
            "ok": True,
            "options": [
                {
                    "slug": "linkedin-post",
                    "label": "LinkedIn Post",
                    "description": "Track a LinkedIn post with date, company, and URL.",
                    "href": "/search/posts/new/",
                },
                {
                    "slug": "project",
                    "label": "Project",
                    "description": "Add a project timeline and delivery notes.",
                    "href": "/search/projects/new/",
                },
                {
                    "slug": "blog-post",
                    "label": "Blog",
                    "description": "Store blog posts from Medium, Dev.to, or your own site.",
                    "href": "/search/blogs/new/",
                },
                {
                    "slug": "research-paper",
                    "label": "Research Paper",
                    "description": "Track papers with publication source and abstract notes.",
                    "href": "/search/research/new/",
                },
            ],
        }
    )


@login_required
def switch_mode_view(request):
    if request.user.is_staff:
        messages.info(request, "Admin accounts always stay in admin mode.")
        return redirect("searcher:organization")

    messages.error(request, "Only admin users can access admin mode.")
    return redirect("searcher:home")


@login_required
def profile_view(request):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    profile, _created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={
            "full_name": request.user.get_full_name() or request.user.username,
            "role_title": "Team Member",
            "linkedin_profile_url": "https://www.linkedin.com/in/",
        },
    )

    if request.method == "POST":
        form = UserProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully.")
            return redirect("searcher:profile")
    else:
        form = UserProfileForm(instance=profile)

    return render(request, "searcher/profile.html", {"form": form, "profile": profile})


@login_required
def post_list_view(request):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    filter_form = PostFilterForm(request.GET or None)
    posts = LinkedInPost.objects.filter(user=request.user)
    posts = _apply_post_filters(posts, filter_form)

    paginator = Paginator(posts.order_by("-posted_at", "-id"), 15)
    page_number = request.GET.get("page")
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    query_params = request.GET.copy()
    query_params.pop("page", None)
    querystring = query_params.urlencode()

    return render(
        request,
        "searcher/posts_list.html",
        {"page_obj": page_obj, "posts": page_obj.object_list, "filter_form": filter_form, "querystring": querystring},
    )


@login_required
def post_create_view(request):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    profile, _created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={
            "full_name": request.user.get_full_name() or request.user.username,
            "role_title": "Team Member",
            "linkedin_profile_url": "https://www.linkedin.com/in/",
        },
    )

    if request.method == "POST":
        form = LinkedInPostForm(request.POST)
        if form.is_valid():
            post = form.save(commit=False)
            post.user = request.user
            post.profile = profile
            post.save()
            messages.success(request, "Post created successfully.")
            return redirect("searcher:posts")
    else:
        form = LinkedInPostForm()

    return render(request, "searcher/post_form.html", {"form": form, "is_edit": False})


@login_required
def post_update_view(request, post_id: int):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    post = get_object_or_404(LinkedInPost, id=post_id, user=request.user)

    if request.method == "POST":
        form = LinkedInPostForm(request.POST, instance=post)
        if form.is_valid():
            form.save()
            messages.success(request, "Post updated successfully.")
            return redirect("searcher:posts")
    else:
        form = LinkedInPostForm(instance=post)

    return render(request, "searcher/post_form.html", {"form": form, "is_edit": True})


@login_required
def post_delete_view(request, post_id: int):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    post = get_object_or_404(LinkedInPost, id=post_id, user=request.user)
    if request.method == "POST":
        post.delete()
        messages.success(request, "Post deleted.")
        return redirect("searcher:posts")

    return render(request, "searcher/post_delete.html", {"post": post})


@login_required
def project_list_view(request):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    projects = Project.objects.filter(user=request.user)
    return render(request, "searcher/projects_list.html", {"projects": projects})


@login_required
def project_create_view(request):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    if request.method == "POST":
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.user = request.user
            project.save()
            messages.success(request, "Project created successfully.")
            return redirect("searcher:projects")
    else:
        form = ProjectForm()

    return render(request, "searcher/project_form.html", {"form": form, "is_edit": False})


@login_required
def project_update_view(request, project_id: int):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    project = get_object_or_404(Project, id=project_id, user=request.user)
    if request.method == "POST":
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, "Project updated successfully.")
            return redirect("searcher:projects")
    else:
        form = ProjectForm(instance=project)

    return render(request, "searcher/project_form.html", {"form": form, "is_edit": True})


@login_required
def project_delete_view(request, project_id: int):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    project = get_object_or_404(Project, id=project_id, user=request.user)
    if request.method == "POST":
        project.delete()
        messages.success(request, "Project deleted.")
        return redirect("searcher:projects")

    return render(request, "searcher/project_delete.html", {"project": project})


@login_required
def blog_list_view(request):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    blogs = BlogPost.objects.filter(user=request.user)
    return render(request, "searcher/blogs_list.html", {"blogs": blogs})


@login_required
def blog_create_view(request):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    if request.method == "POST":
        form = BlogPostForm(request.POST)
        if form.is_valid():
            blog = form.save(commit=False)
            blog.user = request.user
            blog.save()
            messages.success(request, "Blog entry created successfully.")
            return redirect("searcher:blogs")
    else:
        form = BlogPostForm()

    return render(request, "searcher/blog_form.html", {"form": form, "is_edit": False})


@login_required
def blog_update_view(request, blog_id: int):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    blog = get_object_or_404(BlogPost, id=blog_id, user=request.user)
    if request.method == "POST":
        form = BlogPostForm(request.POST, instance=blog)
        if form.is_valid():
            form.save()
            messages.success(request, "Blog entry updated successfully.")
            return redirect("searcher:blogs")
    else:
        form = BlogPostForm(instance=blog)

    return render(request, "searcher/blog_form.html", {"form": form, "is_edit": True})


@login_required
def blog_delete_view(request, blog_id: int):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    blog = get_object_or_404(BlogPost, id=blog_id, user=request.user)
    if request.method == "POST":
        blog.delete()
        messages.success(request, "Blog entry deleted.")
        return redirect("searcher:blogs")

    return render(request, "searcher/blog_delete.html", {"blog": blog})


@login_required
def research_paper_list_view(request):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    papers = ResearchPaper.objects.filter(user=request.user)
    return render(request, "searcher/research_list.html", {"papers": papers})


@login_required
def research_paper_create_view(request):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    if request.method == "POST":
        form = ResearchPaperForm(request.POST)
        if form.is_valid():
            paper = form.save(commit=False)
            paper.user = request.user
            paper.save()
            messages.success(request, "Research paper created successfully.")
            return redirect("searcher:research")
    else:
        form = ResearchPaperForm()

    return render(request, "searcher/research_form.html", {"form": form, "is_edit": False})


@login_required
def research_paper_update_view(request, paper_id: int):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    paper = get_object_or_404(ResearchPaper, id=paper_id, user=request.user)
    if request.method == "POST":
        form = ResearchPaperForm(request.POST, instance=paper)
        if form.is_valid():
            form.save()
            messages.success(request, "Research paper updated successfully.")
            return redirect("searcher:research")
    else:
        form = ResearchPaperForm(instance=paper)

    return render(request, "searcher/research_form.html", {"form": form, "is_edit": True})


@login_required
def research_paper_delete_view(request, paper_id: int):
    blocked = _require_user_mode(request)
    if blocked:
        return blocked

    paper = get_object_or_404(ResearchPaper, id=paper_id, user=request.user)
    if request.method == "POST":
        paper.delete()
        messages.success(request, "Research paper deleted.")
        return redirect("searcher:research")

    return render(request, "searcher/research_delete.html", {"paper": paper})


@user_passes_test(lambda u: u.is_staff)
def organization_view(request):
    filter_form = PostFilterForm(request.GET or None)
    posts = LinkedInPost.objects.select_related("user", "profile")
    posts = _apply_post_filters(posts, filter_form)

    today = timezone.localdate()
    user_summary = (
        posts.values("user", "user__username")
        .annotate(
            total_posts=Count("id"),
            last_7_days_posts=Count("id", filter=Q(posted_at__date__gte=today - timedelta(days=7))),
        )
        .order_by("-total_posts", "user__username")
    )

    company_summary = posts.values("company_name").annotate(post_count=Count("id")).order_by("-post_count", "company_name")

    paginator = Paginator(posts.order_by("-posted_at", "-id"), 25)
    page_number = request.GET.get("page")
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    query_params = request.GET.copy()
    query_params.pop("page", None)
    querystring = query_params.urlencode()

    return render(
        request,
        "searcher/organization.html",
        {
            "posts": page_obj.object_list,
            "page_obj": page_obj,
            "querystring": querystring,
            "filter_form": filter_form,
            "total_posts": posts.count(),
            "total_users": posts.values("user").distinct().count(),
            "user_summary": user_summary,
            "company_summary": company_summary,
        },
    )


@user_passes_test(lambda u: u.is_staff)
def export_posts_csv_view(request):
    filter_form = PostFilterForm(request.GET or None)
    posts = LinkedInPost.objects.select_related("user", "profile")
    posts = _apply_post_filters(posts, filter_form)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="organization_posts.csv"'

    writer = csv.writer(response)
    writer.writerow(["username", "full_name", "company", "post_title", "posted_at", "post_url"])

    for post in posts:
        writer.writerow([
            post.user.username,
            post.profile.full_name if post.profile else "",
            post.company_name,
            post.post_title,
            post.posted_at.isoformat() if post.posted_at else "",
            post.post_url,
        ])

    return response


def _write_xlsx_response(filename: str, workbook: Workbook) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@user_passes_test(lambda u: u.is_staff)
def export_posts_xlsx_view(request):
    filter_form = PostFilterForm(request.GET or None)
    posts = LinkedInPost.objects.select_related("user", "profile")
    posts = _apply_post_filters(posts, filter_form)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Organization Posts"
    sheet.append(["username", "full_name", "company", "post_title", "posted_at", "post_url"])

    for post in posts:
        sheet.append(
            [
                post.user.username,
                post.profile.full_name if post.profile else "",
                post.company_name,
                post.post_title,
                post.posted_at.isoformat() if post.posted_at else "",
                post.post_url,
            ]
        )

    return _write_xlsx_response("organization_posts.xlsx", workbook)


@user_passes_test(lambda u: u.is_staff)
def export_user_summary_csv_view(request):
    filter_form = PostFilterForm(request.GET or None)
    posts = LinkedInPost.objects.select_related("user")
    posts = _apply_post_filters(posts, filter_form)

    today = timezone.localdate()
    summary = (
        posts.values("user__username")
        .annotate(
            total_posts=Count("id"),
            last_7_days_posts=Count("id", filter=Q(posted_at__date__gte=today - timedelta(days=7))),
        )
        .order_by("-total_posts", "user__username")
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="user_post_summary.csv"'

    writer = csv.writer(response)
    writer.writerow(["username", "total_posts", "last_7_days_posts"])

    for row in summary:
        writer.writerow([row["user__username"], row["total_posts"], row["last_7_days_posts"]])

    return response


@user_passes_test(lambda u: u.is_staff)
def export_user_summary_xlsx_view(request):
    filter_form = PostFilterForm(request.GET or None)
    posts = LinkedInPost.objects.select_related("user")
    posts = _apply_post_filters(posts, filter_form)

    today = timezone.localdate()
    summary = (
        posts.values("user__username")
        .annotate(
            total_posts=Count("id"),
            last_7_days_posts=Count("id", filter=Q(posted_at__date__gte=today - timedelta(days=7))),
        )
        .order_by("-total_posts", "user__username")
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "User Summary"
    sheet.append(["username", "total_posts", "last_7_days_posts"])
    for row in summary:
        sheet.append([row["user__username"], row["total_posts"], row["last_7_days_posts"]])

    return _write_xlsx_response("user_post_summary.xlsx", workbook)


@user_passes_test(lambda u: u.is_staff)
def export_single_user_posts_csv_view(request, user_id: int):
    user_model = get_user_model()
    target_user = get_object_or_404(user_model, id=user_id)

    query_data = request.GET.copy()
    query_data["user_name"] = target_user.username
    filter_form = PostFilterForm(query_data or None)

    posts = LinkedInPost.objects.select_related("user", "profile").filter(user=target_user)
    posts = _apply_post_filters(posts, filter_form)

    post_count = posts.count()
    link_count = posts.exclude(post_url="").count()
    first_post = posts.order_by("posted_at").first()
    last_post = posts.order_by("-posted_at").first()

    date_from = first_post.posted_at.isoformat() if first_post and first_post.posted_at else ""
    date_to = last_post.posted_at.isoformat() if last_post and last_post.posted_at else ""
    full_name = (
        (getattr(first_post.profile, "full_name", "") if first_post and first_post.profile else "")
        or target_user.get_full_name()
        or target_user.username
    )

    safe_username = slugify(target_user.username) or f"user-{target_user.id}"
    filename = f"{safe_username}_posts_report.csv"

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(["report", "single_user_posts"])
    writer.writerow(["username", target_user.username])
    writer.writerow(["full_name", full_name])
    writer.writerow(["post_count", post_count])
    writer.writerow(["link_count", link_count])
    writer.writerow(["date_from", date_from])
    writer.writerow(["date_to", date_to])
    writer.writerow(["exported_at", timezone.now().isoformat()])
    writer.writerow([])

    writer.writerow(["post_title", "company_name", "posted_at", "post_url", "notes"])
    for post in posts:
        writer.writerow(
            [
                post.post_title,
                post.company_name,
                post.posted_at.isoformat() if post.posted_at else "",
                post.post_url,
                post.notes,
            ]
        )

    return response


@user_passes_test(lambda u: u.is_staff)
def export_single_user_posts_xlsx_view(request, user_id: int):
    user_model = get_user_model()
    target_user = get_object_or_404(user_model, id=user_id)

    query_data = request.GET.copy()
    query_data["user_name"] = target_user.username
    filter_form = PostFilterForm(query_data or None)

    posts = LinkedInPost.objects.select_related("user", "profile").filter(user=target_user)
    posts = _apply_post_filters(posts, filter_form)

    post_count = posts.count()
    link_count = posts.exclude(post_url="").count()
    first_post = posts.order_by("posted_at").first()
    last_post = posts.order_by("-posted_at").first()

    date_from = first_post.posted_at.isoformat() if first_post and first_post.posted_at else ""
    date_to = last_post.posted_at.isoformat() if last_post and last_post.posted_at else ""
    full_name = (
        (getattr(first_post.profile, "full_name", "") if first_post and first_post.profile else "")
        or target_user.get_full_name()
        or target_user.username
    )

    safe_username = slugify(target_user.username) or f"user-{target_user.id}"
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["field", "value"])
    summary_sheet.append(["username", target_user.username])
    summary_sheet.append(["full_name", full_name])
    summary_sheet.append(["post_count", post_count])
    summary_sheet.append(["link_count", link_count])
    summary_sheet.append(["date_from", date_from])
    summary_sheet.append(["date_to", date_to])
    summary_sheet.append(["exported_at", timezone.now().isoformat()])

    posts_sheet = workbook.create_sheet(title="Posts")
    posts_sheet.append(["post_title", "company_name", "posted_at", "post_url", "notes"])
    for post in posts:
        posts_sheet.append(
            [
                post.post_title,
                post.company_name,
                post.posted_at.isoformat() if post.posted_at else "",
                post.post_url,
                post.notes,
            ]
        )

    return _write_xlsx_response(f"{safe_username}_posts_report.xlsx", workbook)
